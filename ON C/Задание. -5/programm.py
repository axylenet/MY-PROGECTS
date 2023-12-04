from openpyxl import load_workbook, Workbook #Гаврюшкин 21ИС-21
from os.path import join, abspath

class NotAllData(Exception): #проверка, просмотрели файл, но нужные данные отсутствуют
    pass

data_path = join(".", "in.txt.xlsx") #указываем путь к файлу
data_path = abspath(data_path) #абсолютный путь, возвращает значения от моей папки

wb = load_workbook(filename=data_path, data_only=True, read_only=True) #фунция-загрузка книги и работа с файлами и формулами

wsn = list(wb.sheetnames) #список листов

print(wsn)

wsdate = None #даты нет

for i in wsn: #цикл проверки книги на столбец "Фамилия читателя"
    if wb[i]["B1"].value == "Фамилия читателя":
        wsdate = i
if wsdate == None:
    raise NotAllData("No data with Фамилия читателя")

ws = wb[wsdate] #присваиваю значение листа, ссылаясь на wsdate (для обработки)

shapka = [cell.value for cell in next(
    ws.iter_rows(min_row=1, min_col=1, max_row=1, max_col=ws.max_column))] #итератор проходка по колонкам и строкам,все значения в shapka для дальнейшей обработки (мин. строка, мин. колонка,макс. строка, макс.колонка)

chit = {} #словарь

for row in ws.iter_rows(min_row=2, min_col=1, max_row=ws.max_row, max_col=ws.max_column):
    if len(row) > 0:
        fam = row[1].value
        if fam is not None:
            famdata = [cell.value for cell in row]
            if fam not in chit:
                chit[fam] = []
            chit[fam].append(famdata) #Чтобы вытащить все фамилия читателей

for fam in chit:
    print(f"Фамилия читателя: {fam}, кол-во визитов в библиотеку: {len(chit[fam])}")

wb.close #закрываем

#создание отчёта по каждому читателю

for fam in chit:
    exname, *_ = fam.split()
    wb = Workbook()
    ws = wb.active
    ws.title = "Фамилия читателей"

    ws.append(shapka)
    for row in chit[fam]:
        ws.append(row)

    # Собираем путь где сохранить файл
    exfilname = join(".", "out.text",(exname + ".xlsx"))
    exfilname = abspath(exfilname)

    print(exfilname)

    wb.save(exfilname)
    wb.close

    print('\nВсе данные из исходного файла обработаны.')
    print('Файлы сформированы и сохранены в каталог')
